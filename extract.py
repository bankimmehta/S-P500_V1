#!/usr/bin/env python3
"""
extract.py — pull S&P 500 from yfinance -> Excel + cached JSON. Resumable,
permanent cache, paced/backoff (see client.py). Mirrors the NSE tool's outputs:
ReadMe, Overview, Analysis, Flags, Income, Balance, Cash Flow, KeyMetrics.

  python extract.py --limit 25      # TEST first
  python extract.py                 # full S&P 500 (~503; first run ~15-25 min)
  python extract.py --only AAPL,MSFT,JPM
  python extract.py --refresh
  python extract.py --rebuild       # rebuild Excel from cache, no fetching
"""

import argparse
import json
import sys
import datetime

import pandas as pd

import config
import client
import parse
import analysis
import universe

STMT_TABS = ["Income", "Balance", "Cash Flow"]


def _write_excel(path, overview, stmts, metrics, analysis_df, flags_df, meta):
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    with pd.ExcelWriter(path, engine="openpyxl") as xl:
        pd.DataFrame(list(meta.items()), columns=["field", "value"]).to_excel(
            xl, sheet_name="ReadMe", index=False)
        overview.to_excel(xl, sheet_name="Overview", index=False)
        analysis_df.to_excel(xl, sheet_name="Analysis", index=False)
        flags_df.to_excel(xl, sheet_name="Flags", index=False)
        for tab in STMT_TABS:
            (stmts[tab] if not stmts[tab].empty
             else pd.DataFrame({"info": ["(no data)"]})).to_excel(xl, sheet_name=tab, index=False)
        metrics.to_excel(xl, sheet_name="KeyMetrics", index=False)
        for ws in xl.sheets.values():
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="2F5496")
                cell.alignment = Alignment(horizontal="center")
            ws.freeze_panes = "A2"
            for i, _ in enumerate(ws[1], start=1):
                ws.column_dimensions[get_column_letter(i)].width = 22 if i <= 3 else 14


def build_workbook(records, meta_by_ticker, logfn=print):
    """Assemble Excel + JSON from {ticker: record}. Shared by extract & build_dataset."""
    config.OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    items = [(tk, rec) for tk, rec in records.items() if client.is_valid(rec)]
    if not items:
        logfn("Nothing valid to write.")
        return None
    sec_med = analysis.sector_medians(dict(items), meta_by_ticker)

    overview = pd.DataFrame([parse.snapshot(rec, meta_by_ticker.get(tk, {})) for tk, rec in items])
    analysis_df = pd.DataFrame([analysis.summary_row(rec, tk, meta_by_ticker.get(tk, {}), sec_med)
                                for tk, rec in items])
    flag_rows = []
    for tk, rec in items:
        flag_rows.extend(analysis.flags_long(rec, tk, meta_by_ticker.get(tk, {})))
    flags_df = pd.DataFrame(flag_rows) if flag_rows else pd.DataFrame(
        columns=["ticker", "type", "flag", "note"])

    stmt_tables = {}
    for tab in STMT_TABS:
        frames = []
        for tk, rec in items:
            stmts, _ = parse.statements(rec)
            df = stmts.get(tab)
            if df is None or df.empty:
                continue
            d = df.reset_index()
            d.insert(0, "Ticker", tk)
            d.insert(1, "Company", (rec.get("info") or {}).get("longName") or tk)
            frames.append(d)
        stmt_tables[tab] = pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()
    metrics = pd.concat([parse.metrics_long(rec) for _, rec in items], ignore_index=True)

    meta = {
        "source": "yfinance (Yahoo Finance)", "universe": "S&P 500",
        "generated": datetime.datetime.now().strftime("%Y-%m-%d %H:%M"),
        "stocks": len(items),
        "note": "USD. Revenue growth & all analysis from annual statements. "
                "Valuation compares each stock to its GICS-sector median P/E & P/B "
                "across the loaded universe. Statement years are columns.",
    }
    xlsx = config.OUTPUT_DIR / config.XLSX_NAME
    _write_excel(xlsx, overview, stmt_tables, metrics, analysis_df, flags_df, meta)
    json.dump({tk: rec for tk, rec in items}, open(config.OUTPUT_DIR / config.DATA_JSON, "w", encoding="utf-8"))
    logfn(f"Wrote {xlsx}")
    logfn(f"Wrote {config.OUTPUT_DIR / config.DATA_JSON} (for the app)")
    n_inv = int(analysis_df["investigate"].sum()) if not analysis_df.empty else 0
    logfn(f"Analysis: {n_inv}/{len(items)} flagged 'investigate'")
    return xlsx


def main():
    ap = argparse.ArgumentParser(description="S&P 500 yfinance extractor")
    ap.add_argument("--limit", type=int, default=None)
    ap.add_argument("--only", default=None)
    ap.add_argument("--refresh", action="store_true")
    ap.add_argument("--rebuild", action="store_true", help="rebuild from cache, no fetching")
    args = ap.parse_args()

    config.LOG_DIR.mkdir(exist_ok=True)
    stamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    client.configure_logging(config.LOG_DIR / f"extract_{stamp}.log")
    log = client.log

    rows = universe.load()
    meta_by_ticker = {r["ticker"]: r for r in rows}
    if args.only:
        want = {s.strip().upper() for s in args.only.split(",")}
        rows = [r for r in rows if r["ticker"] in want]
    elif args.limit:
        rows = rows[:args.limit]

    records, summary = {}, []
    if args.rebuild:
        log.info("Rebuild: reading cache only.")
        for r in rows:
            p = client._cache_path(r["ticker"])
            if p.exists():
                try:
                    records[r["ticker"]] = json.load(open(p, encoding="utf-8"))
                    summary.append((r["ticker"], "OK (cache)"))
                except Exception as e:  # noqa: BLE001
                    summary.append((r["ticker"], f"BAD CACHE: {e}"))
    else:
        total = len(rows)
        log.info("Fetching %d stocks (permanent cache, resumable, ~%.1f-%.1fs/stock).",
                 total, config.REQUEST_DELAY, config.REQUEST_DELAY + config.REQUEST_JITTER)
        start = datetime.datetime.now()
        for i, r in enumerate(rows, 1):
            tk = r["ticker"]
            rec = client.get_stock(tk, refresh=args.refresh)
            if rec is not None and client.is_valid(rec):
                records[tk] = rec
                summary.append((tk, "OK"))
            else:
                summary.append((tk, "EMPTY/FAILED"))
            if i % 25 == 0 or i == total:
                el = (datetime.datetime.now() - start).total_seconds()
                eta = datetime.timedelta(seconds=int(el / i * (total - i)))
                log.info(">> %d/%d · %d API calls · ETA ~%s", i, total, client.calls_made(), eta)

    ok = [t for t, s in summary if s.startswith("OK")]
    bad = [t for t, s in summary if not s.startswith("OK")]
    log.info("DONE. OK: %d | EMPTY/FAILED: %d | API calls: %d", len(ok), len(bad), client.calls_made())
    if bad:
        log.warning("Failed/empty (re-run to retry): %s", ", ".join(bad[:40]) + (" ..." if len(bad) > 40 else ""))
    if not records:
        log.error("No usable records."); sys.exit(1)

    build_workbook(records, meta_by_ticker, log.info)
    log.info("Tabs: ReadMe, Overview, Analysis, Flags, Income, Balance, Cash Flow, KeyMetrics")


if __name__ == "__main__":
    main()
